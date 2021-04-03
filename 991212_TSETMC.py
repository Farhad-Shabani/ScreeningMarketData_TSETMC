import re
from openpyxl import workbook
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup


Url0 = 'http://www.tsetmc.com/Loader.aspx?ParTree=15'
Url1 = 'http://www.tsetmc.com/tsev2/data/instinfofast.aspx?i='
Url2 = 'http://www.tsetmc.com/loader.aspx?ParTree=151311&i='

Namad = [['3863538898378476','&c=70%20'],['25357135030606405','&c=67%20'],
        ['47302318535715632','&c=57%20'],['17528249960294496','&c=56%20'],
        ['46982154647719707','&c=56%20'],['6757220448540984','&c=53%20'],
        ['15472396110662150','&c=53%20'],['33808206014018431','&c=53%20'],
        ['34890845654517313','&c=53%20'],['3493306453706327','&c=44%20'],
        ['11622051128546106','&c=44%20'],['40611478183231802','&c=44%20'],
        ['44153164692325703','&c=44%20'],['64298008532791199','&c=44%20'],
        ['36899214178084525','&c=43%20'],['12746730665870442','&c=42%20'],
        ['52975109254504632','&c=42%20'],['35964395659427029','&c=38%20'],
        ['63380098535169030','&c=38%20'],['67030488744129337','&c=38%20'],
        ['46752599569017089','&c=31%20'],['23214828924506640','&c=25%20'],
        ['28450080638096732','&c=01%20'],['33420285433308219','&c=01%20'],
        ['60783654574662426','&c=90%20'],['24662567615903665','&c=65%20'],
        ['3050342257199174','&c=38%20'],
        ['59866041653103343','&c=66%20'],['60654872678917533','&c=53%20'],
        ['44802346787824971','&c=53%20'],['7711282667602555','&c=44%20'],
        ['37281199178613855','&c=53%20'],['13937270451301973','&c=42%20'],
        ['48010225447410247','&c=39%20'],['2400322364771558','&c=39%20'],
        ['26014913469567886','&c=39%20'],['52232388263291380','&c=39%20'],
        ['37204371816016200','&c=39%20']]

Result0 = requests.get(Url0)
Src0 = Result0.content
Soup1 = BeautifulSoup(Src0, 'lxml')

# To Scrape Index of Bourse -----------------------------------------------------------
TseTmc1 = Soup1.find_all('tbody')[0].find_all('td')
Index1 = float(''.join(TseTmc1[3].text.split()[0].split(',')))
Index1C = ''.join(TseTmc1[3].text.split()[1].split(','))
Index2 = float(''.join(TseTmc1[5].text.split()[0].split(',')))
Index2C = ''.join(TseTmc1[5].text.split()[1].split(','))
TotalValue1 = int(int(''.join(TseTmc1[13].div['title'].split(',')))/1000000)

# To Scrape Index of FaraBourse --------------------------------------------------------
TseTmc2 = Soup1.find_all('tbody')[25].find_all('td')
Index3 = float(''.join(TseTmc2[3].text.split()[0].split(',')))
Index3C = ''.join(TseTmc2[3].text.split()[1].split(','))
TotalValue2 = int(int(''.join(TseTmc2[11].div['title'].split(',')))/1000000)

Row = list()
DataSet0 = list()
DataSet = list()

for i in range(len(Namad)):
    RevUrl1 = Url1 + Namad[i][0] + Namad[i][1]
    Result1 = requests.get(RevUrl1)
    Src1 = Result1.content
    Split = str(Src1).split(';')
    TradeInfo = Split[0].split(',')
    OrderBook = Split[2].split(',')
    if len(OrderBook) > 1:
        OrderBook1 = OrderBook[0].split('@')
        OrderBook2 = OrderBook[1].split('@')
        OrderBook3 = OrderBook[2].split('@')
        StakeHolder = Split[4].split(',')
    else:
        OrderBook1 = OrderBook2 = OrderBook3 = [0,0,0]
    print(OrderBook1)

    # To calculate Parameters ---------------------------------------------------------
    TradeVolume = round(int(TradeInfo[9])/1000000,2)
    TradeValue = int(int(TradeInfo[10])/1000000)
    HighPrice = int(TradeInfo[6])
    LowPrice = int(TradeInfo[7])
    MarkPrice = int(TradeInfo[3])
    MarkPriceP = round((int(TradeInfo[3])/int(TradeInfo[5]))-1,4)
    LastPrice = int(TradeInfo[2])
    LastPriceP = round((int(TradeInfo[2])/int(TradeInfo[5]))-1,4)

    # To calculate Average of buy & Sell per capita -----------------------------------
    if len(StakeHolder) > 5 and int(StakeHolder[5]) > 0:
        BuyCapita = int((int(StakeHolder[0])/int(StakeHolder[5]))*LastPrice/1000000)
    else: BuyCapita = 0
    if len(StakeHolder) > 8 and int(StakeHolder[8]) > 0:
        SellCapita = int((int(StakeHolder[3])/int(StakeHolder[8]))*LastPrice/1000000)
    else: SellCapita = 0
    if len(StakeHolder) > 0 and SellCapita > 0:
        CapitaRatio = round(BuyCapita/SellCapita,2) 
    else: CapitaRatio = 0
    
    # To calculate Capial -- BaseVol -- QTotTran5JAvg -- P/E ---------------------
    RevUrl2 = Url2 + Namad[i][0]
    Result2 = requests.get(RevUrl2)
    Src2 = Result2.content
    Soup2 = BeautifulSoup(Src2, 'lxml')
    MainPage = Soup2.body.find_all('script')[1].string.split(',')
    Name = MainPage[12].split('=')[1][1:-1]
    EstimatedEPS = MainPage[9]
    EstimatedEPS = int(re.findall("\\d+",EstimatedEPS)[0])
    Capital = MainPage[10]
    Capital = int(int(re.findall("\\d+",Capital)[0])/1000)
    BaseVol = MainPage[8]
    BaseVol = int(re.findall("\\d+",BaseVol)[0])
    QTotTran5JAvg = int(MainPage[24].split('=')[1][1:-1])
    PtoE = round(MarkPrice / EstimatedEPS,2)
    HighValid = MainPage[16]
    HighValid = int(re.findall("\\d+",HighValid)[0])
    LowValid = MainPage[17]
    LowValid = int(re.findall("\\d+",LowValid)[0])
    if QTotTran5JAvg > 0: TradeRatio = TradeVolume/QTotTran5JAvg *1000000
    else: TradeRatio = 0

    # To Build database as a list -----------------------------------------------
    Row.append(Name)                 # 0
    Row.append(Capital)              # 1
    Row.append(BaseVol)              # 2
    Row.append(QTotTran5JAvg)        # 3
    Row.append(TradeVolume)          # 4
    Row.append(TradeRatio)           # 5
    Row.append(TradeValue)           # 6
    Row.append(HighValid)            # 7 
    Row.append(LowValid)             # 8 
    Row.append(MarkPrice)            # 9
    Row.append(MarkPriceP)           # 10
    Row.append(LastPrice)            # 11
    Row.append(LastPriceP)           # 12
    Row.append(EstimatedEPS)         # 13
    Row.append(PtoE)                 # 14
    Row.append(int(OrderBook1[0]))   # 15
    Row.append(int(OrderBook1[1]))   # 16
    Row.append(int(OrderBook1[2]))   # 17
    Row.append(int(OrderBook1[3]))   # 18
    Row.append(int(OrderBook1[4]))   # 19
    Row.append(int(OrderBook1[5]))   # 20
    Row.append(BuyCapita)            # 21
    Row.append(SellCapita)           # 22
    Row.append(CapitaRatio)          # 23

    DataSet0.append(Row)
    DataSet = DataSet0[:34]
    Row = []

# Number of Namads ---------------------------------------------------------
HighCount, PositiveCount, NegativeCount, LowCount = 0, 0, 0, 0
BuyQueueRow = list()
BuyQueue = list()
SellQueueRow = list() 
SellQueue = list()

for i in DataSet:  
    if i[11] == i[7]:
        HighCount += 1
        BuyQueueValue = int(i[11] * i[16] / 1000000)
        BuyQueueRow.append(i[0])                     # 0
        BuyQueueRow.append(i[11])                    # 1
        BuyQueueRow.append(round(i[16]/1000000,1))   # 2
        BuyQueueRow.append(BuyQueueValue)            # 3
        BuyQueue.append(BuyQueueRow)
        BuyQueueRow = []
    elif i[11] != i[7] and i[12] >= 0: PositiveCount +=1
    elif i[11] != i[8] and i[12] < 0: NegativeCount +=1
    elif i[11] == i[8]:
        LowCount += 1
        SellQueueValue = int(i[11] * i[19] / 1000000)
        SellQueueRow.append(i[0])                     # 0
        SellQueueRow.append(i[11])                    # 1
        SellQueueRow.append(round(i[19]/1000000,1))   # 2
        SellQueueRow.append(SellQueueValue)           # 3
        SellQueue.append(SellQueueRow)
        SellQueueRow = []

# Sorting --------------------------------------------------------------------

VolRatioSort = sorted(DataSet,key = lambda x: x[5],reverse=True)  
ValueSort = sorted(DataSet,key = lambda x: x[6],reverse=True)
BuyCapitaSort = sorted(DataSet,key = lambda x: x[-3],reverse=True)
SellCapitaSort = sorted(DataSet,key = lambda x: x[-2],reverse=True)  
BuyQueue.sort(key = lambda x: x[3], reverse=True) 
SellQueue.sort(key = lambda x: x[3], reverse=True) 


print(ValueSort)
print(BuyQueue)   
print(SellQueue)
print(HighCount,PositiveCount,NegativeCount,LowCount)  


wb = load_workbook(filename='DailyReport.xlsx')
ws = wb.active
# Index --------------------------------------------------------------------
ws['C7'] = Index1
ws['D7'] = Index1C
ws['C8'] = Index2
ws['D8'] = Index2C
ws['C9'] = Index3
ws['D9'] = Index3C
ws['J7'] = TotalValue1
ws['J8'] = TotalValue2

TotalValueVabank = 0
for i in ValueSort:
    TotalValueVabank += i[6]
ws['J11'] = TotalValueVabank

Num = 0
BaseVolNum = 0
for i in DataSet:
    Num += 1
    if i[4] >= i[2]/1000000:
        BaseVolNum +=1
ws['J12'] = round(BaseVolNum / Num, 2)



# Number of Namads --------------------------------------------------------
ws['P12'] = HighCount
ws['O12'] = PositiveCount
ws['N12'] = NegativeCount
ws['M12'] = LowCount

# Important Namads --------------------------------------------------------
ws['B19'] = DataSet0[33][11]
ws['B20'] = DataSet0[33][12]
ws['C19'] = DataSet0[34][11]
ws['C20'] = DataSet0[34][12]
ws['D19'] = DataSet0[35][11]
ws['D20'] = DataSet0[35][12]
ws['E19'] = DataSet0[36][11]
ws['E20'] = DataSet0[36][12]
ws['G19'] = DataSet0[37][11]
ws['G20'] = DataSet0[37][12]
ws['H19'] = DataSet0[32][11]
ws['H20'] = DataSet0[32][12]
ws['I19'] = DataSet0[31][11]
ws['I20'] = DataSet0[31][12]
ws['J19'] = DataSet0[30][11]
ws['J20'] = DataSet0[30][12]

# ValueSort ---------------------------------------------------------------
for j in range(10):
    ws['M{}'.format(25+j)] = ValueSort[j][0]
    ws['N{}'.format(25+j)] = ValueSort[j][6]

# Queue -------------------------------------------------------------------
if len(BuyQueue) > 7: LBuy = 7
else: LBuy = len(BuyQueue)
for z in range(LBuy):
    ws['B{}'.format(37+z)] = BuyQueue[z][0]
    ws['C{}'.format(37+z)] = BuyQueue[z][1]
    ws['D{}'.format(37+z)] = BuyQueue[z][2]
    ws['E{}'.format(37+z)] = BuyQueue[z][3]

if len(SellQueue) > 7: LSell = 7
else: LSell = len(SellQueue)
for t in range(LSell):
    ws['G{}'.format(37+t)] = SellQueue[t][0]
    ws['H{}'.format(37+t)] = SellQueue[t][1]
    ws['I{}'.format(37+t)] = SellQueue[t][2]
    ws['J{}'.format(37+t)] = SellQueue[t][3]

# Suspicious In & Out -----------------------------------------------------
if len(BuyCapitaSort) > 5: BCSort = 5
else: BCSort = len(BuyCapitaSort)
for k in range(BCSort):
    ws['B{}'.format(49+k)] = BuyCapitaSort[k][0]
    ws['D{}'.format(49+k)] = BuyCapitaSort[k][-3]

if len(VolRatioSort) > 5: VRSort = 5
else: VRSort = len(VolRatioSort)
for l in range(VRSort):
    ws['E{}'.format(49+l)] = VolRatioSort[l][0]
    ws['G{}'.format(49+l)] = VolRatioSort[l][5]

if len(SellCapitaSort) > 5: SCSort = 5
else: SCSort = len(SellCapitaSort)
for m in range(SCSort):
    ws['H{}'.format(49+m)] = SellCapitaSort[m][0]
    ws['J{}'.format(49+m)] = SellCapitaSort[m][-2]

wb.save(filename="DailyReport.xlsx")