import re, time, requests
from bs4 import BeautifulSoup
from openpyxl import workbook
from openpyxl import load_workbook
from src.Scrape_Index import Scrape_Index
from src.Database_Maker import Database_Maker, Sorted_Database


def Export_Database(portfolio,Selected_Stocks):
    StartTime = time.time()
    Index = Scrape_Index()
    Portfolio_Database = Database_Maker(portfolio)
    Selected_Stocks_Database = Database_Maker(Selected_Stocks)
    CountCache, VolRatioSort, ValueSort, BuyCapitaSort, SellCapitaSort, BuyQueue, SellQueue = Sorted_Database(Portfolio_Database,portfolio)

    wb = load_workbook(filename='Dashboard_Template.xlsx')
    ws = wb.active
    
    # Index ---------------------------------------------------------------------------------------------------------

    ws['C7'] = Index[0]
    ws['D7'] = Index[1]
    ws['C8'] = Index[2]
    ws['D8'] = Index[3]
    ws['C9'] = Index[4]
    ws['D9'] = Index[5]
    ws['J7'] = Index[6]
    ws['J8'] = Index[7]

    TotalValueVabank = 0
    for i in ValueSort:
        TotalValueVabank += i[8]
    ws['J11'] = TotalValueVabank

    Num = 0
    BaseVolNum = 0
    for i in Portfolio_Database:
        Num += 1
        if i[7] >= i[2]/1000000:
           BaseVolNum +=1
    ws['J12'] = round(BaseVolNum / Num, 2)


    # Number of Stocks -----------------------------------------------------------------------------------------------

    ws['P12'] = CountCache[0]
    ws['O12'] = CountCache[1]
    ws['N12'] = CountCache[2]
    ws['M12'] = CountCache[3]
    ws['E41'] = CountCache[4]
    ws['J41'] = CountCache[5]    
    ws['D52'] = CountCache[6]
    ws['J52'] = CountCache[7]


    # To Show price of Selected Stocks -------------------------------------------------------------------------------
    Selected_Stocks_Map = {0:'B', 1:'C', 2:'D', 3:'E', 4:'G', 5:'H', 6:'I', 7:'J'}
    for i in Selected_Stocks_Map:
        ws['{}15'.format(Selected_Stocks_Map[i])] = Selected_Stocks_Database[i][0]  
        ws['{}16'.format(Selected_Stocks_Map[i])] = Selected_Stocks_Database[i][11]
        ws['{}17'.format(Selected_Stocks_Map[i])] = Selected_Stocks_Database[i][12]


    # ValueSort -----------------------------------------------------------------------------------------------------
    for j in range(10):
        ws['M{}'.format(22+j)] = ValueSort[j][0]
        ws['N{}'.format(22+j)] = ValueSort[j][8]


    # Queue ---------------------------------------------------------------------------------------------------------
    if len(BuyQueue) > 7: LBuy = 7
    else: LBuy = len(BuyQueue)
    for z in range(LBuy):
        ws['B{}'.format(34+z)] = BuyQueue[z][0]
        ws['C{}'.format(34+z)] = BuyQueue[z][1]
        ws['D{}'.format(34+z)] = BuyQueue[z][2]
        ws['E{}'.format(34+z)] = BuyQueue[z][3]

    if len(SellQueue) > 7: LSell = 7
    else: LSell = len(SellQueue)
    for t in range(LSell):
        ws['G{}'.format(34+t)] = SellQueue[t][0]
        ws['H{}'.format(34+t)] = SellQueue[t][1]
        ws['I{}'.format(34+t)] = SellQueue[t][2]
        ws['J{}'.format(34+t)] = SellQueue[t][3]


    # Suspicious In & Out ----------------------------------------------------------------------------------------=--
    if len(BuyCapitaSort) > 5: BCSort = 5
    else: BCSort = len(BuyCapitaSort)
    for k in range(BCSort):
        ws['B{}'.format(47+k)] = BuyCapitaSort[k][0]
        ws['D{}'.format(47+k)] = BuyCapitaSort[k][-5]

    if len(VolRatioSort) > 5: VRSort = 5
    else: VRSort = len(VolRatioSort)
    for l in range(VRSort):
        ws['E{}'.format(47+l)] = VolRatioSort[l][0]
        ws['G{}'.format(47+l)] = VolRatioSort[l][22]

    if len(SellCapitaSort) > 5: SCSort = 5
    else: SCSort = len(SellCapitaSort)
    for m in range(SCSort):
        ws['H{}'.format(47+m)] = SellCapitaSort[m][0]
        ws['J{}'.format(47+m)] = SellCapitaSort[m][-4]
    
    EndTime = time.time()
    print('Your Dashboard is ready! \nIt took {} seconds to create the output excel.'.format(int(EndTime - StartTime)))

    return wb.save(filename="TSETMC_DailyDashboard.xlsx")


# Give a list of 7 your desired stocks in order to make the Excel-based Dashboard --------------------------------------

Portfolio = {'ثمسکن': ['3863538898378476','&c=70%20'], 
             'تنوین': ['25357135030606405','&c=67%20'],
             'ونوین': ['47302318535715632','&c=57%20'],
             'وتوسم': ['17528249960294496','&c=56%20'],
             'وصنا': ['46982154647719707','&c=56%20'],
             'سشمال': ['6757220448540984','&c=53%20'],
             'سكرما': ['15472396110662150','&c=53%20'],
             'سمازن': ['33808206014018431','&c=53%20'],
             'ساربيل': ['34890845654517313','&c=53%20'],
             'شوينده': ['3493306453706327','&c=44%20'],
             'شپاكسا': ['11622051128546106','&c=44%20'],
             'شدوص': ['40611478183231802','&c=44%20'],
             'شگل': ['44153164692325703','&c=44%20'],
             'ساينا': ['64298008532791199','&c=44%20'],
             'شفا': ['36899214178084525','&c=43%20'],
             'بهپاك': ['12746730665870442','&c=42%20'],
             'غمارگ': ['52975109254504632','&c=42%20'],
             'قشكر': ['35964395659427029','&c=38%20'],
             'قنيشا': ['63380098535169030','&c=38%20'],
             'قپيرا': ['67030488744129337','&c=38%20'],
             'بترانس': ['46752599569017089','&c=31%20'],
             'پكرمان': ['23214828924506640','&c=25%20'],
             'سيمرغ': ['28450080638096732','&c=01%20'],
             'زپارس': ['33420285433308219','&c=01%20'],
             'وهنر': ['60783654574662426','&c=90%20'],
             'وسنا': ['24662567615903665','&c=65%20'],
             'قنقش': ['3050342257199174','&c=38%20'],
             'نوين': ['59866041653103343','&c=66%20'],
             'سقاين': ['60654872678917533','&c=53%20'],
             'ساروج': ['44802346787824971','&c=53%20'],
             'شاراك': ['7711282667602555','&c=44%20'],
             'سيدكو': ['37281199178613855','&c=53%20'],
             'وبشهر': ['13937270451301973','&c=42%20'],
             'وبانك': ['48010225447410247','&c=39%20'],
             'سنوین': ['36995197800118822', '&c=56%20']

}

# Give a list of your 8 desired stocks to show the real-time prices in your dashboard ----------------------------------

Selected_Stocks = {
             'وبانك': ['48010225447410247','&c=39%20'],
             'شستا': ['2400322364771558','&c=39%20'],
             'وغدير': ['26014913469567886','&c=39%20'],
             'واميد': ['52232388263291380','&c=39%20'],
             'وصندوق': ['37204371816016200','&c=39%20'],
             'وبشهر': ['13937270451301973','&c=42%20'],
             'سيدكو': ['37281199178613855','&c=53%20'],
             'شاراك': ['7711282667602555','&c=44%20']
}


Export_Database(Portfolio, Selected_Stocks)